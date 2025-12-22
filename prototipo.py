import pyautogui
import pyperclip
import time
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from pathlib import Path
import pandas as pd
import shutil
import os

TIMEOUT = 60  # segundos

downloads = Path("D:/Users/mchci/Downloads")

# def esperar_qualquer_download(timeout=60):
#     inicio = time.time()
#     arquivos_iniciais = set(os.listdir(downloads))
#     for arquivo in downloads.iterdir():
#             if arquivo.is_file():
#                 # data de modificação
#                 data_arquivo = date.fromtimestamp(arquivo.stat().st_mtime)

#                 if data_arquivo == hoje:
#                     if not arquivo.endswith('.crdownload'):
#                         return os.path.join(downloads, arquivo)
#                     if time.time() - inicio > timeout:
#                         raise TimeoutError('⏰ Nenhum download detectado')
#                     time.sleep(1)

def esperar_qualquer_download(timeout=60):
    inicio = time.time()
    arquivos_iniciais = set(os.listdir(downloads))

    while True:
        arquivos_atuais = set(os.listdir(downloads))
        novos = arquivos_atuais - arquivos_iniciais

        for arquivo in novos:
            if not arquivo.endswith('.crdownload'):
                return os.path.join(downloads, arquivo)
        
        if time.time() - inicio > timeout:
            raise TimeoutError('⏰ Nenhum download detectado')

# -------- FUNCAO PARA CSV ---------
def adicionar_cpf(df, cpf):
    # garantir string
    cpf = str(cpf).strip()

    linhas_vazias = df[df['CPF'] == '']

    if linhas_vazias.empty:
        print('⚠️ Nenhuma linha disponível para inserir CPF')
        return df  # ⛔ PARA AQUI

    idx = linhas_vazias.index[-1]
    df.at[idx, 'CPF'] = cpf
    return df
# ---------------------------
pyautogui.FAILSAFE = True

data_atual = datetime.today()


dia = data_atual.strftime("%d")
mes = data_atual.month
ano = data_atual.strftime("%Y")

meses = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro"
}

mes_nome = meses[mes]

print(dia, mes_nome, ano)

#inserir o numero de registros
num = int(input("Digite o número de registros: "))

def minimizar():
    pyautogui.moveTo(pyautogui.moveTo(x=1806, y=7))
    pyautogui.click()
    time.sleep(0.5)

#minimizar vscode
minimizar()
time.sleep(0.5)

#POSIÇÕES
#POSICAO INICIAL LINHA SIGAMA = 336

posicao_i_nome_S = [77, 332]
posicao_a_nome_S = posicao_i_nome_S

# posicao_i_nome_E = [252, 309]
# posicao_a_nome_E = posicao_i_nome_E

posicao_i_cpf_S = [839, 332]
posicao_a_cpf_S = posicao_i_cpf_S

# posicao_i_cpf_E = [590, 309]
# posicao_a_cpf_E = posicao_i_cpf_E

posicao_i_lupa = [1690, 332]
posicao_a_lupa = posicao_i_lupa

for j in range(num):

    #copiar nome - SIGAMA
    pyautogui.moveTo(posicao_a_nome_S)
    pyautogui.tripleClick()
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(0.5)

    # --------- USANDO EXCEL ---------------
    # arquivo = f"C:\SIGAMA\Documentos Solicitaçoes de Acesso\\{ano}\\{mes}\\{dia}\Controlde de Solicitação.xlsx"
    # #coluna = 'A'  # coluna desejada
    # nome = pyperclip.paste()

    # wb = load_workbook(arquivo)
    # ws = wb["Controle de Solicitação"] # ou wb["NomeDaAba"]

    # linha = 1

    # # encontra a primeira célula vazia
    # while ws[f"{coluna}{linha}"].value is not None:
    #     linha += 1

    # # escreve o nome
    # ws[f"{coluna}{linha}"] = nome

    # # salva
    # wb.save(arquivo)

    # ---------------------------------------

    # ------------ USANDO CSV (TESTE) --------------

    nome = pyperclip.paste()
    arquivo = Path(
    "C:/SIGAMA/Documentos Solicitaçoes de Acesso",
    str(ano),
    str(mes_nome),
    str(dia),
    "teste.csv"
)

    df = pd.read_csv(arquivo, sep=';', dtype=str)
    df = df.fillna('')
    df.loc[len(df)] = {
        'NOME': nome,
        'CPF': ''
    }
    df.to_csv(arquivo, sep=';', index=False)

    #copiar cpf SIGAMA
    pyautogui.moveTo(posicao_a_cpf_S)
    pyautogui.doubleClick()
    pyautogui.hotkey('ctrl', 'c')
    cpf = pyperclip.paste()
    # #cola cpf no excel
    # coluna = 'B'

    # # escreve o cpf
    # ws[f"{coluna}{linha}"] = cpf

    # # salva
    # wb.save(arquivo)

    # colunas = ['A', 'B']   # colunas de destino

    # nome_cpf = [ws[f"{col}{linha}"].value for col in colunas]
    # nome_cpf_tratado = nome_cpf.replace("\t", " - ")

    # Path(f"C:\SIGAMA\Documentos Solicitaçoes de Acesso\\{ano}\\{mes}\\{dia}", nome_cpf_tratado).mkdir(exist_ok=True)

    # ------------------------ COPIANDO CPF CSV ------------------
    df = pd.read_csv(arquivo, sep=';', dtype=str).fillna('')

    df = adicionar_cpf(df, cpf)

    df.to_csv(arquivo, sep=';', index=False)

    # --------------------------------------------------------
    
    #localizar lupa
    pyautogui.moveTo(posicao_a_lupa)
    pyautogui.click()
    time.sleep(1.2)

    #localizar documentos
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('./Automacao/image/anexo_image.png', confidence = 0.8))
    x1, y1 = pyautogui.locateCenterOnScreen('./Automacao/image/anexo_image.png', confidence = 0.8)

    #clicar no doc
    pyautogui.moveTo(x1, y1+27)

    for i in range(5):
        y1 = y1+23
        pyautogui.moveTo(x1, y1)
        time.sleep(0.5)
        pyautogui.click()
        time.sleep(0.2)

    time.sleep(2)

    pyautogui.click(pyautogui.locateCenterOnScreen('./Automacao/image/X_image.png', confidence = 0.8))

    # --------------- COPIANDO NOME E CPF CSV -------------------

    ultima_linha = df.iloc[-1]
    nome = ultima_linha['NOME']
    cpf = ultima_linha['CPF']

    nome_cpf_tratado = f"{nome} - {cpf}"

    # -----------------------------------------------------------

    # downloads = Path.home() / "Downloads"

    destino = Path(
    "C:/SIGAMA/Documentos Solicitaçoes de Acesso",
    str(ano),
    str(mes_nome),
    str(dia),
    nome_cpf_tratado
)
    
    destino.mkdir(exist_ok=True)

    hoje = date.today()
    agora = datetime.now()

    for arquivo in downloads.iterdir():
        if arquivo.is_file():
            if arquivo.suffix != '.crdownload':
                    data_arquivo = datetime.fromtimestamp(arquivo.stat().st_mtime)
                    if agora - data_arquivo < timedelta(seconds=15):
                        time.sleep(4)
                        shutil.move(arquivo, destino / arquivo.name)
    # #localizar chrome
    # pyautogui.click(pyautogui.locateCenterOnScreen('.\image\chrome_image.png', confidence = 0.8))


    posicao_a_nome_S[1] = posicao_a_nome_S[1] + 40

    # posicao_a_nome_E[1] = posicao_a_nome_E[1] + 26

    posicao_a_cpf_S[1] = posicao_a_cpf_S[1] + 40

    # posicao_a_cpf_E[1] = posicao_a_cpf_E[1] + 26

    posicao_a_lupa[1] = posicao_a_lupa[1] + 40

caminho_csv = Path(
    "C:/SIGAMA/Documentos Solicitaçoes de Acesso",
    str(ano),
    str(mes_nome),
    str(dia),
    "teste.csv"
)

caminho_pasta = Path(
    "C:/SIGAMA/Documentos Solicitaçoes de Acesso",
    str(ano),
    str(mes_nome),
    str(dia)
)

os.startfile(caminho_csv)
os.startfile(caminho_pasta)