#Importando bibliotecas
import pyautogui
import pyperclip
import time
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from pathlib import Path
import pandas as pd
import shutil
import os
import sys
from tkinter import simpledialog, Tk

#Importando utilitários
from utils.connection import checar_conectividade
from utils.validate import validate_excel, validate_folders

status = checar_conectividade()

if status != "OK":
    print(f"Problema de conexão: {status}")
    sys.exit()

validate_excel()

TIMEOUT = 60

def resource_path(relative_path):
    """ Retorna o caminho correto para PyInstaller ou execução normal """
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

files_directory = Path.home() / "Downloads"

validate_folders(files_directory)

def esperar_qualquer_download(timeout=60):
    inicio = time.time()
    arquivos_iniciais = set(os.listdir(files_directory))

    while True:
        arquivos_atuais = set(os.listdir(files_directory))
        novos = arquivos_atuais - arquivos_iniciais

        for arquivo in novos:
            if not arquivo.endswith('.crdownload'):
                return os.path.join(files_directory, arquivo)
        
        if time.time() - inicio > timeout:
            raise TimeoutError('Nenhum download detectado')

pyautogui.FAILSAFE = True

data_atual = datetime.today()

dia = data_atual.strftime("%d")
mes = data_atual.month
ano = data_atual.strftime("%Y")

meses = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro"
}

mes_nome = meses[mes]

print(dia, mes_nome, ano)

#inserir o numero de registros

root = Tk()
root.withdraw()

num = simpledialog.askinteger(
    "SIGAMA",
    "Digite o número de registros:"
)

if num is None:
    sys.exit()
#num = int(input("Digite o número de registros: "))

def minimizar():
    pyautogui.moveTo(pyautogui.moveTo(x=1806, y=7))
    pyautogui.click()
    time.sleep(0.5)

#minimizar vscode
#minimizar()
#time.sleep(0.5)

#POSIÇÕES
#POSICAO INICIAL LINHA SIGAMA = 336

img_nome = resource_path("image/nome_image.jpg")

PNx, PNy = pyautogui.locateCenterOnScreen(img_nome, confidence= 0.6)
posicao_i_nome_S = [PNx, PNy + 40]
posicao_a_nome_S = posicao_i_nome_S

# posicao_i_nome_E = [252, 309]
# posicao_a_nome_E = posicao_i_nome_E

img_cpf = resource_path("image/cpf_image.jpg")

PCx, PCy = pyautogui.locateCenterOnScreen(img_cpf, confidence= 0.6)
posicao_i_cpf_S = [PCx, PCy + 40]
posicao_a_cpf_S = posicao_i_cpf_S

# posicao_i_cpf_E = [590, 309]
# posicao_a_cpf_E = posicao_i_cpf_E

img_operacoes = resource_path("image/operacoes_image.jpeg")
PLx, PLy = pyautogui.locateCenterOnScreen(img_operacoes, confidence= 0.5)
posicao_i_lupa = [PLx + 52, PLy + 34]
posicao_a_lupa = posicao_i_lupa

#Criar pastas

base = Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso")
pasta = base / ano / mes_nome / dia
pasta.mkdir(parents=True, exist_ok=True)

origem = Path("Z:/SIGAMA/Controle de Solicitação.xlsx")
destino_pasta = Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso",str(ano), str(mes_nome), str(dia))

destino_arquivo = destino_pasta / origem.name

if not destino_arquivo.exists():
    shutil.copy(origem, destino_pasta)
    print("Arquivo copiado com sucesso")
else:
    print("Arquivo já existe, não foi copiado")

for j in range(num):

    status = checar_conectividade()

    if status != "OK":
        print(f"Problema de conexão: {status}")
        sys.exit()

    #copiar nome - SIGAMA
    pyautogui.moveTo(posicao_a_nome_S)
    pyautogui.tripleClick()
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(0.5)

    arquivo = Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso",str(ano), str(mes_nome), str(dia), "Controle de Solicitação.xlsx")
    coluna = 'A'  # coluna desejada
    nome = pyperclip.paste()

    wb = load_workbook(arquivo)
    ws = wb["Controle de Solicitação"] # ou wb["NomeDaAba"]

    linha = 1

    # encontra a primeira célula vazia
    while ws[f"{coluna}{linha}"].value is not None:
        linha += 1

    # escreve o nome
    ws[f"{coluna}{linha}"] = nome

    # salva
    wb.save(arquivo)

    #copiar cpf SIGAMA
    pyautogui.moveTo(posicao_a_cpf_S)
    pyautogui.doubleClick()
    pyautogui.hotkey('ctrl', 'c')
    cpf = pyperclip.paste()

    #cola cpf no excel
    coluna = 'B'

    # escreve o cpf
    ws[f"{coluna}{linha}"] = cpf

    # salva
    wb.save(arquivo)

    colunas = ['A', 'B']   # colunas de destino

    #criar pasta com nome e cpf
    nome_cpf = [ws[f"{col}{linha}"].value for col in colunas]
    nome_cpf_tratado = f"{nome} - {cpf}"

    Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso", str(ano), str(mes_nome), str(dia), nome_cpf_tratado).mkdir(exist_ok=True)

    #localizar lupa
    pyautogui.moveTo(posicao_a_lupa)
    pyautogui.click()
    
    status = checar_conectividade()

    if status != "OK":
        print(f"Problema de conexão: {status}")
        sys.exit()

    time.sleep(0.2)

    #Localizar anexo de documentos
    img_anexo = resource_path("image/anexo_image.jpg")
    pyautogui.moveTo(pyautogui.locateCenterOnScreen(img_anexo, confidence = 0.8))
    x1, y1 = pyautogui.locateCenterOnScreen(img_anexo, confidence = 0.8)
    x1 = x1 - 25

    #clicar nos documentos
    pyautogui.moveTo(x1, y1+27)
    for i in range(5):
        y1 = y1+23
        pyautogui.moveTo(x1, y1)
        time.sleep(0.5)
        pyautogui.click()
        time.sleep(0.2)
    time.sleep(0.2)
    
    img_X = resource_path("image/X_image.jpg")
    pyautogui.click(pyautogui.locateCenterOnScreen(img_X, confidence = 0.8))

    #Colocar documentos na pasta
    destino = Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso",
    str(ano),
    str(mes_nome),
    str(dia),
    nome_cpf_tratado
)
    
    hoje = date.today()
    agora = datetime.now()

    #identificar erro no diretório de arquivos
    for arquivo in files_directory.iterdir():
        if arquivo.is_file():
            if arquivo.suffix != '.crdownload':
                data_arquivo = datetime.fromtimestamp(arquivo.stat().st_mtime)
                if agora - data_arquivo < timedelta(seconds=15):
                    time.sleep(0.2)
                    shutil.move(arquivo, destino / arquivo.name)
            else:
                print("Erro no download")
                sys.exit()

    #iteração das posições
    posicao_a_nome_S[1] = posicao_a_nome_S[1] + 40

    posicao_a_cpf_S[1] = posicao_a_cpf_S[1] + 40

    posicao_a_lupa[1] = posicao_a_lupa[1] + 40

#Abrir paginas no final da execução
caminho_pasta = Path(
    "Z:/SIGAMA/Documentos Solicitaçoes de Acesso",
    str(ano),
    str(mes_nome),
    str(dia)
)

caminho_excel = Path("Z:/SIGAMA/Documentos Solicitaçoes de Acesso",str(ano), str(mes_nome), str(dia), "Controle de Solicitação.xlsx")

# os.startfile(caminho_csv)
os.startfile(caminho_pasta)
os.startfile(caminho_excel)